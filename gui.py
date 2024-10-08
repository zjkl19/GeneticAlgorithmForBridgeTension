import tkinter as tk
from tkinter import ttk
from tkinter import scrolledtext
from tkinter import messagebox
from tkinter import filedialog
from genetic_algorithm import run_genetic_algorithm
import numpy as np
import json
import shutil
import time
from tkinter import ttk  # 导入ttk模块，用于创建Notebook
from openpyxl import Workbook
from datetime import datetime
import logging

# 配置日志记录器
logging.basicConfig(filename='calculation_log.txt', level=logging.INFO,
                    format='%(asctime)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')

import openpyxl
from openpyxl import Workbook

def generate_excel_with_timestamp(initial_forces, target_forces, influence_matrix, best_individual, adjusted_individual):
    # 获取当前时间戳
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"计算结果_{timestamp}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "计算结果"

    # 设置表头
    ws["A1"] = "初始索力"
    ws["B1"] = "目标索力"
    # 影响矩阵的表头将根据矩阵的列数动态设置
    n = len(initial_forces)  # 索力数量
    matrix_start_col = 3  # 影响矩阵开始的列编号
    for j in range(matrix_start_col, matrix_start_col + n):
        ws.cell(row=1, column=j).value = f"影响因子{j-2}"

    matrix_end_col = matrix_start_col + n - 1
    ws[openpyxl.utils.get_column_letter(matrix_end_col + 1) + "1"] = "原始最优解"
    ws[openpyxl.utils.get_column_letter(matrix_end_col + 2) + "1"] = "调整阈值后的最优解"
    ws[openpyxl.utils.get_column_letter(matrix_end_col + 3) + "1"] = "原始最终索力"
    ws[openpyxl.utils.get_column_letter(matrix_end_col + 4) + "1"] = "原始误差%"
    ws[openpyxl.utils.get_column_letter(matrix_end_col + 5) + "1"] = "调整后最终索力"
    ws[openpyxl.utils.get_column_letter(matrix_end_col + 6) + "1"] = "调整后误差%"


    # 填充初始索力和目标索力
    for i, (init_force, target_force) in enumerate(zip(initial_forces, target_forces), start=2):
        ws[f"A{i}"] = init_force
        ws[f"B{i}"] = target_force

    # 填充影响矩阵
    for i, row in enumerate(influence_matrix, start=2):
        for j, val in enumerate(row, start=3):  # 从第3列开始填充
            ws.cell(row=i, column=j, value=val)

    n = len(initial_forces)  # 索力数量
    matrix_start_col = 3  # 影响矩阵开始的列编号
    matrix_end_col = matrix_start_col + n - 1

    # 填充最优解和调整阈值后的最优解
    for i, (best_val, adj_val) in enumerate(zip(best_individual, adjusted_individual), start=2):
        ws.cell(row=i, column=matrix_end_col + 1, value=best_val)
        ws.cell(row=i, column=matrix_end_col + 2, value=adj_val)

    best_solution_col = openpyxl.utils.get_column_letter(matrix_end_col + 1)
    adjusted_solution_col = openpyxl.utils.get_column_letter(matrix_end_col + 2)

    # 对每个索力的调整后结果和误差百分比进行计算
    # 注意：这里直接将最优解和调整后最优解作为列向量参与运算
    for i in range(2, n + 2):
        # 对于每个索力，计算调整后的结果
        final_force_formula = f'=SUM('
        adjusted_final_force_formula = f'=SUM('
        for j in range(matrix_start_col, matrix_end_col + 1):
            final_force_formula += f'{openpyxl.utils.get_column_letter(j)}{i}*${best_solution_col}${j-1}+'
            adjusted_final_force_formula += f'{openpyxl.utils.get_column_letter(j)}{i}*${adjusted_solution_col}${j-1}+'
        final_force_formula = final_force_formula.rstrip('+') + f')+$A{i}'
        adjusted_final_force_formula = adjusted_final_force_formula.rstrip('+') + f')+$A{i}'

        # 将计算结果公式填充到单元格中
        ws.cell(row=i, column=matrix_end_col + 3, value=final_force_formula)
        ws.cell(row=i, column=matrix_end_col + 4, value=f'=ROUND(({openpyxl.utils.get_column_letter(matrix_end_col + 3)}{i}-$B{i})/$B{i}*100, 2)')    
        ws.cell(row=i, column=matrix_end_col + 5, value=adjusted_final_force_formula)
        ws.cell(row=i, column=matrix_end_col + 6, value=f'=ROUND(({openpyxl.utils.get_column_letter(matrix_end_col + 5)}{i}-$B{i})/$B{i}*100, 2)')

    # 保存Excel文件
    wb.save(filename)
    return filename  # 返回生成的文件名，以便后续操作（如提示用户文件已保存等）)
import os
import glob


class App:
    def __init__(self, root):
        self.root = root
        self.root.title("索力调整遗传算法")
        self.selected_metric = tk.StringVar(value="最大误差比例")  # 默认为最大误差比例

        # 创建一个外框架
        outer_frame = ttk.Frame(self.root)
        outer_frame.pack(fill="both", expand=True)

        # 创建Canvas和滚动条
        self.canvas = tk.Canvas(outer_frame)
        self.scrollbar = ttk.Scrollbar(outer_frame, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = ttk.Frame(self.canvas)

        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.canvas.configure(
                scrollregion=self.canvas.bbox("all")
            )
        )

        self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=self.scrollbar.set)

        self.canvas.pack(side="left", fill="both", expand=True)
        self.scrollbar.pack(side="right", fill="y")
        # 在滚动框架内创建所有控件
        self.create_widgets()

    def create_widgets(self):
        row_index = 0  # 控制布局的行号
        tk.Label(self.scrollable_frame, text="种群大小:").pack()
        self.population_size = tk.Entry(self.scrollable_frame)
        self.population_size.pack()
        row_index += 1

        tk.Label(self.scrollable_frame, text="交叉率:").pack()
        self.crossover_rate = tk.Entry(self.scrollable_frame)
        self.crossover_rate.pack()
        row_index += 1

        tk.Label(self.scrollable_frame, text="变异率:").pack()
        self.mutation_rate = tk.Entry(self.scrollable_frame)
        self.mutation_rate.pack()
        row_index += 1

        tk.Label(self.scrollable_frame, text="最小调整力(例：-300):").pack()
        self.min_adjustment = tk.Entry(self.scrollable_frame)
        self.min_adjustment.pack()
        row_index += 1

        tk.Label(self.scrollable_frame, text="最大调整力(例：300):").pack()
        self.max_adjustment = tk.Entry(self.scrollable_frame)
        self.max_adjustment.pack()
        row_index += 1

        tk.Label(self.scrollable_frame, text="索力最大值限制:").pack()
        self.max_force = tk.Entry(self.scrollable_frame)
        self.max_force.pack()
        row_index += 1

        tk.Label(self.scrollable_frame, text="误差控制范围(%):").pack()
        self.tolerance = tk.Entry(self.scrollable_frame)
        self.tolerance.pack()
        row_index += 1

        tk.Label(self.scrollable_frame, text="索下限(可用Tab/逗号/空格分隔):").pack()
        self.lower_bounds_text = tk.Text(self.scrollable_frame, height=2)
        self.lower_bounds_text.pack()
        row_index += 1

        tk.Label(self.scrollable_frame, text="索上限(可用Tab/逗号/空格分隔):").pack()
        self.upper_bounds_text = tk.Text(self.scrollable_frame, height=2)
        self.upper_bounds_text.pack()
        row_index += 1

        # 添加评价指标选择的下拉菜单
        metrics_options = ["最大误差比例", "均方误差"]
        tk.Label(self.scrollable_frame, text="选择评价指标:").pack()
        self.metric_dropdown = ttk.Combobox(self.scrollable_frame, textvariable=self.selected_metric, values=metrics_options, state='readonly')
        self.metric_dropdown.pack()
        row_index += 1

        tk.Label(self.scrollable_frame, text="调整阈值:").pack()
        self.adjustment_threshold = tk.Entry(self.scrollable_frame)
        self.adjustment_threshold.pack()
        self.adjustment_threshold.insert(0, '0')  # 可以设置一个默认阈值，例如0
        row_index += 1

        tk.Label(self.scrollable_frame, text="代数(gen):").pack()
        self.ngen = tk.Entry(self.scrollable_frame)
        self.ngen.pack()
        row_index += 1

        # 添加输入框用于指定运行限制时间（分钟）
        self.time_limit_label = tk.Label(self.scrollable_frame, text="高级搜索运行限制时间（分钟）：")
        self.time_limit_label.pack()
        self.time_limit = tk.Entry(self.scrollable_frame)
        self.time_limit.insert(tk.END, "2")  # 默认值2分钟
        self.time_limit.pack()
        row_index += 2  # 添加一个额外的空行作为分隔

        # initial_forces 输入
        tk.Label(self.scrollable_frame, text="初始索力:").pack()
        self.initial_forces_text = tk.Text(self.scrollable_frame, height=5)
        self.initial_forces_text.pack()
        row_index += 1

        # target_forces 输入
        tk.Label(self.scrollable_frame, text="目标索力:").pack()
        self.target_forces_text = tk.Text(self.scrollable_frame, height=5)
        self.target_forces_text.pack()
        row_index += 1

        # influence_matrix 输入
        tk.Label(self.scrollable_frame, text="影响矩阵:").pack()
        self.influence_matrix_text = tk.Text(self.scrollable_frame, height=5)
        self.influence_matrix_text.pack()
        row_index += 1

        # 添加查看日志按钮
        self.view_log_button = tk.Button(self.scrollable_frame, text="查看日志", command=self.view_log)
        self.view_log_button.pack()
        row_index += 1

        # 创建“打开最新验算Excel”按钮
        self.open_excel_button = tk.Button(self.scrollable_frame, text="打开最新验算Excel", command=self.open_latest_excel)
        self.open_excel_button.pack()
        row_index += 1

        # 添加标签页控件
        self.notebook = ttk.Notebook(self.scrollable_frame)
        self.config_tab = ttk.Frame(self.notebook)  # 配置标签页
        self.notebook.add(self.config_tab, text='试算配置')
        self.notebook.pack(pady=10, fill="x")
        row_index += 1

        # 在试算配置标签页中添加控件
        trial_row_index = 0  # 控制试算配置内的布局行号
        tk.Label(self.config_tab, text="交叉率范围:").pack()
        self.crossover_rate_min = tk.Entry(self.config_tab)
        self.crossover_rate_min.pack(side="left")
        self.crossover_rate_max = tk.Entry(self.config_tab)
        self.crossover_rate_max.pack(side="left")
        trial_row_index += 1

        tk.Label(self.config_tab, text="变异率范围:").pack()
        self.mutation_rate_min = tk.Entry(self.config_tab)
        self.mutation_rate_min.pack(side="left")
        self.mutation_rate_max = tk.Entry(self.config_tab)
        self.mutation_rate_max.pack(side="left")
        trial_row_index += 1

        # 试算按钮
        self.trial_run_button = tk.Button(self.config_tab, text="试算", command=self.trial_run)
        self.trial_run_button.pack()

        # 配置操作按钮
        self.save_config_button = tk.Button(self.scrollable_frame, text="保存配置", command=self.save_config)
        self.save_config_button.pack(side="left")

        self.load_config_button = tk.Button(self.scrollable_frame, text="加载配置", command=self.load_config)
        self.load_config_button.pack(side="left")

        self.backup_config_button = tk.Button(self.scrollable_frame, text="备份配置", command=self.backup_config)
        self.backup_config_button.pack(side="left")
        row_index += 2  # 添加一个额外的空行作为分隔

        # 结果显示区域，位于所有控件的最下方
        self.result_text = tk.Text(self.scrollable_frame, height=10, width=50)
        self.result_text.pack(pady=5)

        # 添加开始计算按钮
        self.start_button = tk.Button(self.scrollable_frame, text="开始计算", command=self.start_calculation)
        self.start_button.pack(side="left")

        # 添加新按钮用于触发高级搜索算法
        self.advanced_search_button = tk.Button(self.scrollable_frame, text="高级搜索", command=self.run_advanced_search)
        self.advanced_search_button.pack(side="left")

        # 设置默认值
        self.population_size.insert(0, '1000')  # 种群大小默认值1000
        self.crossover_rate.insert(0, '0.7')  # 交叉率默认值0.7
        self.mutation_rate.insert(0, '0.2')  # 变异率默认值0.2
        self.min_adjustment.insert(0, '-300')  # 最小调整力默认值-300
        self.max_adjustment.insert(0, '300')  # 最大调整力默认值300
        self.max_force.insert(0, '3000')  # 索力最大值限制默认值3000
        self.tolerance.insert(0, '5')  # 误差控制范围默认值5%
        self.ngen.insert(0, '100')  # 设置代数默认值为100
        self.crossover_rate_min.insert(0, '0.6')
        self.crossover_rate_max.insert(0, '0.9')
        self.mutation_rate_min.insert(0, '0.01')
        self.mutation_rate_max.insert(0, '0.1')
        self.adjustment_threshold.insert(0, '40')
    def open_latest_excel(self):
            # 查找当前目录下所有匹配的Excel文件
            files = glob.glob("计算结果_*.xlsx")
            # 按文件修改时间排序，获取最新的文件
            if files:
                latest_file = max(files, key=os.path.getmtime)
                try:
                    os.startfile(latest_file)  # 尝试使用系统默认方式打开文件
                except Exception as e:
                    messagebox.showerror("错误", f"无法打开文件{latest_file}。\n错误信息: {str(e)}")
            else:
                messagebox.showinfo("提示", "当前目录下没有找到Excel文件，请先生成。")

    def view_log(self):
        import os
        import subprocess
        import sys
        
        log_file_path = 'calculation_log.txt'
        try:
            if sys.platform.startswith('win'):
                os.startfile(log_file_path)
            elif sys.platform.startswith('darwin'):
                subprocess.call(('open', log_file_path))
            else:
                subprocess.call(('xdg-open', log_file_path))
        except Exception as e:
            messagebox.showerror("错误", f"无法打开日志文件: {e}")

    def save_config(self):
        config = {
            "population_size": self.population_size.get(),
            "crossover_rate_min": self.crossover_rate_min.get(),
            "crossover_rate_max": self.crossover_rate_max.get(),
            "mutation_rate_min": self.mutation_rate_min.get(),
            "mutation_rate_max": self.mutation_rate_max.get(),
            "min_adjustment": self.min_adjustment.get(),
            "max_adjustment": self.max_adjustment.get(),
            "max_force": self.max_force.get(),
            "tolerance": self.tolerance.get(),
            "initial_forces": self.initial_forces_text.get("1.0", tk.END).strip(),
            "target_forces": self.target_forces_text.get("1.0", tk.END).strip(),
            "influence_matrix": self.influence_matrix_text.get("1.0", tk.END).strip(),
            "ngen": self.ngen.get(),
            "lower_bounds": self.lower_bounds_text.get("1.0", tk.END).strip(),
            "upper_bounds": self.upper_bounds_text.get("1.0", tk.END).strip(),
            "selected_metric": self.selected_metric.get(),
            "adjustment_threshold": self.adjustment_threshold.get()
        }
        with open("config.json", "w") as f:
            json.dump(config, f, indent=4)

        messagebox.showinfo("配置保存", "配置已成功保存到config.json")
    def load_config(self):
            missing_config = []  # 用于记录缺失的配置项
            try:
                with open("config.json", "r") as f:
                    config = json.load(f)

                def get_config(key, default, widget, is_text=False, is_stringvar=False):
                    if key in config:
                        value = config[key]
                        if is_text:
                            widget.delete("1.0", tk.END)
                            widget.insert("1.0", value)
                        elif is_stringvar:
                            widget.set(value)  # 对于StringVar使用.set()方法
                        else:
                            widget.delete(0, tk.END)
                            widget.insert(0, value)
                    else:
                        missing_config.append(key)
                        if is_text:
                            widget.delete("1.0", tk.END)
                            widget.insert("1.0", default)
                        elif is_stringvar:
                            widget.set(default)  # 对于StringVar使用默认值
                        else:
                            widget.delete(0, tk.END)
                            widget.insert(0, default)

                # 应用基本配置到GUI，使用上面定义的get_config函数
                get_config("population_size", "100", self.population_size)
                get_config("crossover_rate_min", "0.6", self.crossover_rate_min)
                get_config("crossover_rate_max", "0.9", self.crossover_rate_max)
                get_config("mutation_rate_min", "0.01", self.mutation_rate_min)
                get_config("mutation_rate_max", "0.1", self.mutation_rate_max)
                get_config("min_adjustment", "-300", self.min_adjustment)
                get_config("max_adjustment", "300", self.max_adjustment)
                get_config("max_force", "3000", self.max_force)
                get_config("tolerance", "5", self.tolerance)
                get_config("ngen", "100", self.ngen)
                
                # 应用文本输入配置
                get_config("initial_forces", "", self.initial_forces_text, is_text=True)
                get_config("target_forces", "", self.target_forces_text, is_text=True)
                get_config("influence_matrix", "", self.influence_matrix_text, is_text=True)
                
                # 应用lower_bounds、upper_bounds、selected_metric和adjustment_threshold
                get_config("lower_bounds", "-500", self.lower_bounds_text, is_text=True)
                get_config("upper_bounds", "500", self.upper_bounds_text, is_text=True)
                get_config("selected_metric", "最大误差比例", self.selected_metric, is_stringvar=True)
                get_config("adjustment_threshold", "40", self.adjustment_threshold)

                messagebox.showinfo("配置加载", "配置已从config.json加载")
                
                # 如果有缺失的配置项，提醒用户
                if missing_config:
                    missing_items_str = ", ".join(missing_config)
                    messagebox.showwarning("配置缺失", f"以下配置项在文件中缺失，已应用默认值：{missing_items_str}")
            except Exception as e:
                messagebox.showerror("加载配置错误", f"无法加载配置：{e}")


    def backup_config(self):
        try:
            shutil.copy("config.json", "config_backup.json")
            messagebox.showinfo("配置备份", "配置备份已成功创建")
        except Exception as e:
            messagebox.showerror("备份配置错误", f"配置备份失败：{e}")        

    def display_data(self, data, data_type):
        # 向文本区域添加数据
        self.data_display.config(state='normal')  # 允许编辑文本区域来添加数据
        self.data_display.insert(tk.END, f"{data_type}:\n{data}\n")
        self.data_display.config(state='disabled')  # 禁止编辑文本区域
    def validate_input(self):
        # 输入验证函数，确保所有输入参数都有效
        try:
            population_size = int(self.population_size.get())
            crossover_rate = float(self.crossover_rate.get())
            mutation_rate = float(self.mutation_rate.get())
            min_adjustment = float(self.min_adjustment.get())
            max_adjustment = float(self.max_adjustment.get())
            max_force = float(self.max_force.get())
            tolerance = float(self.tolerance.get()) / 100  # 将百分比转换为小数
            ngen = int(self.ngen.get())  # 获取并验证代数的输入

            # 适应多种分隔符，包括制表符、逗号和空格
            lower_bounds_str = self.lower_bounds_text.get("1.0", tk.END).strip()
            upper_bounds_str = self.upper_bounds_text.get("1.0", tk.END).strip()

            # 使用正则表达式匹配多种可能的分隔符
            import re
            lower_bounds = [float(x) for x in re.split(r'[\t, ]+', lower_bounds_str) if x]
            upper_bounds = [float(x) for x in re.split(r'[\t, ]+', upper_bounds_str) if x]

            # 检查维度匹配和逻辑合理性
            assert len(lower_bounds) == len(upper_bounds), "下限和上限的数量必须相同"
            assert all(lb < ub for lb, ub in zip(lower_bounds, upper_bounds)), "每个下限必须小于对应的上限"


            assert population_size > 0, "种群大小必须大于0"
            assert 0 <= crossover_rate <= 1, "交叉率必须在0和1之间"
            assert 0 <= mutation_rate <= 1, "变异率必须在0和1之间"
            assert min_adjustment < max_adjustment, "最小调整力必须小于最大调整力"
            assert max_force > 0, "索力最大值限制必须大于0"
            assert 0 < tolerance < 1, "误差控制范围必须为正值"
            assert ngen > 0, "代数必须大于0"

            return (population_size, crossover_rate, mutation_rate, min_adjustment, max_adjustment, max_force, tolerance, ngen,lower_bounds,upper_bounds)
        except ValueError as e:
            messagebox.showerror("输入错误", "请确保所有输入框都填写了数字。")
            return None
        except AssertionError as e:
            messagebox.showerror("输入错误", str(e))
            return None

    
    def start_calculation(self):
        # 先获取并验证输入参数
        selected_metric = self.selected_metric.get()
        # 将中文标签转换回英文代码
        metric_code = "mse" if selected_metric == "均方误差" else "max_error_percent"

        # 获取输入参数
        params = self.validate_input()
        if params is None:
            return  # 输入验证失败

        # 参数解包
        population_size, crossover_rate, mutation_rate, min_adjustment, max_adjustment, max_force, tolerance, ngen, lower_bounds, upper_bounds = params

        threshold = float(self.adjustment_threshold.get())  # 获取用户输入的阈值

        # 获取和验证初始索力、目标索力和影响矩阵
        try:
            initial_forces_str = self.initial_forces_text.get("1.0", tk.END).strip()
            target_forces_str = self.target_forces_text.get("1.0", tk.END).strip()
            influence_matrix_str = self.influence_matrix_text.get("1.0", tk.END).strip()

            # 将字符串转换为 numpy 数组
            initial_forces = np.fromstring(initial_forces_str, sep='\n')
            target_forces = np.fromstring(target_forces_str, sep='\n')
            influence_matrix = np.fromstring(influence_matrix_str, sep=' ').reshape(-1, initial_forces.size)

            # 确保数据的有效性
            assert initial_forces.size == target_forces.size, "初始索力和目标索力的维度不匹配。"
            assert influence_matrix.shape[0] == influence_matrix.shape[1] == initial_forces.size, "影响矩阵的维度与索力不匹配。"

        except ValueError as e:
            messagebox.showerror("输入校核错误", str(e))
            return  # 终止计算过程
        except Exception as e:
            messagebox.showerror("未知错误", f"在处理输入数据时发生错误: {e}")
            return  # 终止计算过程

        # 运行遗传算法
        try:
            # 在运行遗传算法之前记录起始时间
            start_time = time.time()

            # 调用遗传算法函数，并处理返回值
            best_individual, best_fitness, final_pop, log = run_genetic_algorithm(
                initial_forces, target_forces, influence_matrix, 
                population_size, crossover_rate, mutation_rate, 
                min_adjustment, max_adjustment, max_force, tolerance, ngen, lower_bounds, upper_bounds, selected_metric=metric_code
            )
            best_individual_formatted = [round(x, 0) for x in best_individual]

            # 计算调整后的索力和误差百分比
            adjusted_forces = initial_forces + np.dot(influence_matrix, best_individual)
            adjusted_forces_formatted = [round(x, 0) for x in adjusted_forces]
            # 计算实际误差和误差百分比
            actual_errors = adjusted_forces - target_forces
            original_error_percentages = 100 * actual_errors / target_forces
            original_error_percentages_formatted = [f"{e:.2f}%" for e in original_error_percentages]

            # 找到绝对值误差最大的索
            max_error_idx = np.argmax(np.abs(actual_errors))
            max_error_value = actual_errors[max_error_idx]  # 绝对值误差最大的实际误差值（包括正负号）
            max_error_percentage = round(original_error_percentages[max_error_idx], 2)  # 对应的误差百分比

            # 根据阈值调整索力值
            adjusted_individual = [value if abs(value) > threshold else 0 for value in best_individual]
            adjusted_individual_formatted = [round(x, 0) for x in adjusted_individual]

            # 根据阈值调整索力值后，再次计算误差百分比
            adjusted_forces_after_threshold = initial_forces + np.dot(influence_matrix, adjusted_individual)
            adjusted_forces_after_threshold_formatted = [round(x, 2) for x in adjusted_forces_after_threshold]

            adjusted_error_percentages = 100 * (adjusted_forces_after_threshold - target_forces) / target_forces
            adjusted_error_percentages_formatted = [f"{e:.2f}%" for e in adjusted_error_percentages]

            # 找到调整后的最大误差百分比
            max_adjusted_error_idx = np.argmax(np.abs(adjusted_error_percentages))
            max_adjusted_error_value = adjusted_error_percentages[max_adjusted_error_idx]  # 绝对值误差最大的实际误差值（包括正负号）
            max_adjusted_error_percentage = round(max_adjusted_error_value, 2)  # 对应的误差百分比

            # 检查是否有解满足误差要求
            solution_found = all(abs(e) <= tolerance * 100 for e in adjusted_error_percentages)
            solution_status = f"状态: {'找到解，调整后的最大误差百分比为' if solution_found else f'未找到解，当前计算最接近的解调整后的最大误差百分比为'}: {max_adjusted_error_percentage}%"

            # 记录算法完成后的时间
            end_time = time.time()

            # 计算运行时间
            run_time = end_time - start_time
            # 转换为时分秒
            hours, remainder = divmod(run_time, 3600)
            minutes, seconds = divmod(remainder, 60)

            # 清除旧结果并显示新结果
            self.result_text.delete('1.0', tk.END)
            self.result_text.insert(tk.END, f"{solution_status}\n")
            self.result_text.insert(tk.END, f"最优解的适应度: {best_fitness[0]:.8f}\n")
            self.result_text.insert(tk.END, f"原始最优解: {best_individual_formatted}\n")
            self.result_text.insert(tk.END, f"原始调整后索力: {adjusted_forces_formatted}\n")
            self.result_text.insert(tk.END, "原始各索力的误差百分比: \n" + ", ".join(map(str, original_error_percentages_formatted)) + "\n")
            self.result_text.insert(tk.END, f"绝对值误差最大的是索 {max_error_idx + 1}，误差: {max_error_value:.2f}, 误差百分比: {max_error_percentage}%\n")
            self.result_text.insert(tk.END, f"调整阈值后的最优解: {adjusted_individual_formatted}\n")
            self.result_text.insert(tk.END, f"调整阈值后索力: {adjusted_forces_after_threshold_formatted}\n")
            self.result_text.insert(tk.END, "调整后的误差百分比: \n" + ", ".join(adjusted_error_percentages_formatted) + "\n")
            self.result_text.insert(tk.END, f"运行时间: {int(hours)}时{int(minutes)}分{int(seconds)}秒\n")

            time_str = f"{int(hours)}小时{int(minutes)}分{int(seconds)}秒"

            # 调用 generate_excel_with_timestamp 生成Excel文件
            filename = generate_excel_with_timestamp(
                initial_forces.tolist(), 
                target_forces.tolist(), 
                influence_matrix, 
                best_individual_formatted, 
                adjusted_individual_formatted
            )

            # 构建日志消息
            log_message = (f"计算时间: {time_str}\n"
                        f"最优解的适应度: {best_fitness[0]:.8f}\n"
                        f"原始最优解: {', '.join(map(str, best_individual_formatted))}\n"
                        f"原始各索力的误差百分比: {', '.join(map(lambda x: f'{x:.2f}%', original_error_percentages))}\n"
                        f"绝对值误差最大的是索 {max_error_idx + 1}，误差: {max_error_value:.2f}, 误差百分比: {max_error_percentage:.2f}%\n"
                        f"输入参数: 种群大小={population_size}, 交叉率={crossover_rate}, 变异率={mutation_rate}, "
                        f"最小调整力={min_adjustment}, 最大调整力={max_adjustment}, 索力最大值限制={max_force}, "
                        f"误差控制范围={tolerance*100:.2f}%, 代数={ngen}, "
                        f"下限={', '.join(map(str, lower_bounds))}, 上限={', '.join(map(str, upper_bounds))}, "
                        f"评价指标={selected_metric}, 调整阈值={threshold}\n"
                        f"调整阈值后的最优解: {', '.join(map(str, adjusted_individual_formatted))}\n"
                        f"调整后的各索力的误差百分比: {', '.join(map(lambda x: f'{x:.2f}%', adjusted_error_percentages))}\n"
                        f"{solution_status}\n"
                        "--------------------------------------------------")

            # 记录日志
            logging.info(log_message)

        except Exception as e:
            raise e

    def run_advanced_search(self):
        try:
            time_limit = float(self.time_limit.get()) * 60  # 获取运行限制时间，并转换为秒
        except ValueError:
            messagebox.showerror("输入错误", "请输入有效的运行限制时间（分钟）")
            return

        # 获取并验证输入参数
        selected_metric = self.selected_metric.get()
        metric_code = "mse" if selected_metric == "均方误差" else "max_error_percent"

        params = self.validate_input()
        if params is None:
            return

        population_size, crossover_rate, mutation_rate, min_adjustment, max_adjustment, max_force, tolerance, ngen, lower_bounds, upper_bounds = params
        threshold = float(self.adjustment_threshold.get())

        try:
            initial_forces_str = self.initial_forces_text.get("1.0", tk.END).strip()
            target_forces_str = self.target_forces_text.get("1.0", tk.END).strip()
            influence_matrix_str = self.influence_matrix_text.get("1.0", tk.END).strip()

            initial_forces = np.fromstring(initial_forces_str, sep='\n')
            target_forces = np.fromstring(target_forces_str, sep='\n')
            influence_matrix = np.fromstring(influence_matrix_str, sep=' ').reshape(-1, initial_forces.size)

            assert initial_forces.size == target_forces.size, "初始索力和目标索力的维度不匹配。"
            assert influence_matrix.shape[0] == influence_matrix.shape[1] == initial_forces.size, "影响矩阵的维度与索力不匹配。"

        except ValueError as e:
            messagebox.showerror("输入校核错误", str(e))
            return
        except Exception as e:
            messagebox.showerror("未知错误", f"在处理输入数据时发生错误: {e}")
            return

        try:
            start_time = time.time()
            best_solutions = []
            iteration_count = 0

            while time.time() - start_time < time_limit:
                best_individual, best_fitness, final_pop, log = run_genetic_algorithm(
                    initial_forces, target_forces, influence_matrix,
                    population_size, crossover_rate, mutation_rate,
                    min_adjustment, max_adjustment, max_force, tolerance, ngen, lower_bounds, upper_bounds, selected_metric=metric_code
                )
                
                # 根据阈值调整索力值
                adjusted_individual = [value if abs(value) > threshold else 0 for value in best_individual]

                adjusted_forces_after_threshold = initial_forces + np.dot(influence_matrix, adjusted_individual)
                adjusted_error_percentages = 100 * (adjusted_forces_after_threshold - target_forces) / target_forces
                max_adjusted_error_percentage = max(abs(e) for e in adjusted_error_percentages)

                best_solutions.append((adjusted_individual, max_adjusted_error_percentage))
                iteration_count += 1

                # 记录每次运算的日志
                log_message = (f"第 {iteration_count} 次运算\n"
                            f"当前最优解: {adjusted_individual}\n"
                            f"当前最优解的最大误差百分比: {max_adjusted_error_percentage:.2f}%\n")
                logging.info(log_message)

            best_solutions.sort(key=lambda x: x[1])
            adjusted_individual, max_adjusted_error_percentage = best_solutions[0]
            adjusted_individual_formatted = [round(x, 0) for x in adjusted_individual]

            adjusted_forces_after_threshold = initial_forces + np.dot(influence_matrix, adjusted_individual)
            adjusted_forces_after_threshold_formatted = [round(x, 2) for x in adjusted_forces_after_threshold]

            adjusted_error_percentages = 100 * (adjusted_forces_after_threshold - target_forces) / target_forces
            adjusted_error_percentages_formatted = [f"{e:.2f}%" for e in adjusted_error_percentages]

            solution_found = all(abs(e) <= tolerance * 100 for e in adjusted_error_percentages)
            solution_status = f"状态: {'找到解' if solution_found else f'未找到解，当前计算最接近的解调整后的最大误差百分比为: {max_adjusted_error_percentage:.2f}%'}"

            end_time = time.time()
            run_time = end_time - start_time
            hours, remainder = divmod(run_time, 3600)
            minutes, seconds = divmod(remainder, 60)

            self.result_text.delete('1.0', tk.END)
            self.result_text.insert(tk.END, f"{solution_status}\n")
            self.result_text.insert(tk.END, f"当前最优解的最大误差百分比: {max_adjusted_error_percentage:.2f}%\n")
            self.result_text.insert(tk.END, f"调整阈值后的最优解: {adjusted_individual_formatted}\n")
            self.result_text.insert(tk.END, f"调整阈值后索力: {adjusted_forces_after_threshold_formatted}\n")
            self.result_text.insert(tk.END, "调整后的误差百分比: \n" + ", ".join(adjusted_error_percentages_formatted) + "\n")
            self.result_text.insert(tk.END, f"运行时间: {int(hours)}时{int(minutes)}分{int(seconds)}秒\n")

            time_str = f"{int(hours)}小时{int(minutes)}分{int(seconds)}秒"

            filename = generate_excel_with_timestamp(
                initial_forces.tolist(),
                target_forces.tolist(),
                influence_matrix,
                adjusted_individual_formatted,
                adjusted_individual_formatted
            )

            log_message = (f"计算时间: {time_str}\n"
                        f"当前最优解的最大误差百分比: {max_adjusted_error_percentage:.2f}%\n"
                        f"调整阈值后的最优解: {', '.join(map(str, adjusted_individual_formatted))}\n"
                        f"调整后的各索力的误差百分比: {', '.join(map(lambda x: f'{x:.2f}%', adjusted_error_percentages))}\n"
                        f"{solution_status}\n"
                        "--------------------------------------------------")

            logging.info(log_message)

        except KeyboardInterrupt:
            messagebox.showinfo("中断", "操作已被用户中断")
        except Exception as e:
            raise e

    def trial_run(self):
        try:
            # 从 GUI 获取固定参数
            population_size = int(self.population_size.get())
            min_adjustment = float(self.min_adjustment.get())
            max_adjustment = float(self.max_adjustment.get())
            max_force = float(self.max_force.get())
            tolerance = float(self.tolerance.get()) / 100  # 转换为小数
            ngen = int(self.ngen.get())
            
            # 从文本输入区域获取数据
            initial_forces_str = self.initial_forces_text.get("1.0", tk.END).strip()
            target_forces_str = self.target_forces_text.get("1.0", tk.END).strip()
            influence_matrix_str = self.influence_matrix_text.get("1.0", tk.END).strip()

            # 将字符串转换为 numpy 数组
            initial_forces = np.fromstring(initial_forces_str, sep='\n')
            target_forces = np.fromstring(target_forces_str, sep='\n')
            influence_matrix = np.fromstring(influence_matrix_str, sep=' ').reshape(-1, len(initial_forces))

            # 从 GUI 获取交叉率和变异率的搜索范围
            crossover_rate_min = float(self.crossover_rate_min.get())
            crossover_rate_max = float(self.crossover_rate_max.get())
            mutation_rate_min = float(self.mutation_rate_min.get())
            mutation_rate_max = float(self.mutation_rate_max.get())

            best_fitness = np.inf
            best_params = None

            # 网格搜索
            for crossover_rate in np.arange(crossover_rate_min, crossover_rate_max, 0.1):
                for mutation_rate in np.arange(mutation_rate_min, mutation_rate_max, 0.01):
                    _, fitness, _, _ = run_genetic_algorithm(
                        initial_forces, target_forces, influence_matrix, 
                        population_size, crossover_rate, mutation_rate, 
                        min_adjustment, max_adjustment, max_force, tolerance,ngen
                    )
                    # 确保从元组中提取适应度值
                    fitness = fitness[0]  # 假设适应度值位于元组的第一个位置                   
                    if fitness < best_fitness:
                        best_fitness = fitness
                        best_params = (crossover_rate, mutation_rate)

            # 显示最佳结果和参数
            messagebox.showinfo("试算结果", f"最佳适应度: {best_fitness}\n最佳交叉率: {best_params[0]}\n最佳变异率: {best_params[1]}")
        except Exception as e:
            messagebox.showerror("试算错误", f"试算过程中发生错误: {e}")
def main():
    root = tk.Tk()
    app = App(root)
    root.mainloop()

if __name__ == "__main__":
    main()
